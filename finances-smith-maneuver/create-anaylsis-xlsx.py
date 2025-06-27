from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, numbers

# Create workbook
wb = Workbook()

# ----- Input sheet -----
ws_in = wb.active
ws_in.title = "Input"

inputs = [
    ("Original Mortgage Principal", 600000),
    ("Mortgage Rate (%)", 3.5),
    ("HELOC Rate (%)", 7),
    ("Investment Return (%)", 6),
    ("Marginal Tax Rate (%)", 47),
    ("Amortization (years)", 25),
    ("Payments per Year", 12),
]

for i, (label, value) in enumerate(inputs, start=2):
    ws_in[f"A{i}"] = label
    ws_in[f"B{i}"] = value
    ws_in[f"A{i}"].alignment = Alignment(horizontal="right")
    ws_in[f"A{i}"].font = Font(bold=True)
    ws_in[f"B{i}"].number_format = numbers.FORMAT_NUMBER_00
    ws_in[f"B{i}"].fill = PatternFill(
        "solid", fgColor="FFFFCC"
    )  # light yellow for easy editing

# Set first column width to 40 pixels
ws_in.column_dimensions["A"].width = 40
ws_in.column_dimensions["B"].width = 10

# ----- Scenarios sheet -----
ws_sc = wb.create_sheet("Scenarios")

headers = [
    "Scenario",
    "Mortgage Rate (%)",
    "HELOC Rate (%)",
    "Investment Return (%)",
    "Tax Rate (%)",
    "Monthly Payment",
    "Principal Repaid Year1",
    "HELOC Interest Year1",
    "Portfolio Value Year1",
]

for col, header in enumerate(headers, start=1):
    cell = ws_sc.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    ws_sc.column_dimensions[get_column_letter(col)].width = 20

# Scenario definitions
scenarios = [
    ("Base", "=Input!B3", "=Input!B4", "=Input!B5", "=Input!B6"),
    ("Low Mortgage Rate", "=Input!B3-1", "=Input!B4", "=Input!B5", "=Input!B6"),
    ("High HELOC Rate", "=Input!B3", "=Input!B4+1.5", "=Input!B5", "=Input!B6"),
    ("Bull Market", "=Input!B3", "=Input!B4", "=Input!B5+3", "=Input!B6"),
    ("Bear Market", "=Input!B3", "=Input!B4", "=Input!B5-3", "=Input!B6"),
]

start_row = 2
for idx, scenario in enumerate(scenarios):
    row = start_row + idx
    ws_sc.cell(row=row, column=1, value=scenario[0])
    # columns B-E are formulas or values
    ws_sc.cell(row=row, column=2, value=scenario[1])
    ws_sc.cell(row=row, column=3, value=scenario[2])
    ws_sc.cell(row=row, column=4, value=scenario[3])
    ws_sc.cell(row=row, column=5, value=scenario[4])

    # Monthly Payment formula (column F)
    ws_sc.cell(
        row=row,
        column=6,
        value=f"=PMT(B{row}/100/Input!$B$8, Input!$B$7*Input!$B$8, -Input!$B$2)",
    )

    # Principal Repaid Year1 (column G)
    ws_sc.cell(
        row=row,
        column=7,
        value=f"=-CUMPRINC(B{row}/100/Input!$B$8, Input!$B$7*Input!$B$8, Input!$B$2, 1, 12, 0)",
    )

    # HELOC Interest Year1 (column H)
    ws_sc.cell(
        row=row,
        column=8,
        value=f"=G{row}*C{row}/100",
    )

    # Portfolio Value Year1 (column I)
    ws_sc.cell(
        row=row,
        column=9,
        value=f"=G{row}*(1+D{row}/100)",
    )

# Apply number formats
for row in ws_sc.iter_rows(min_row=2, max_row=6, min_col=2, max_col=9):
    for cell in row:
        if cell.column in (2, 3, 4, 5):  # percent inputs
            cell.number_format = numbers.FORMAT_NUMBER_00
        else:
            cell.number_format = numbers.FORMAT_NUMBER_00

# ----- Chart -----
chart = BarChart()
chart.type = "col"
chart.title = "Portfolio Value After Year 1"
chart.y_axis.title = "Value ($)"
chart.x_axis.title = "Scenario"

data = Reference(ws_sc, min_col=9, min_row=1, max_row=6)  # Portfolio Value Year1
cats = Reference(ws_sc, min_col=1, min_row=2, max_row=6)  # Scenario names
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 10
chart.width = 20

ws_sc.add_chart(chart, "K2")

# ----- Summary sheet -----
ws_sum = wb.create_sheet("Summary")
ws_sum["A2"] = "Highest Portfolio Value (Year 1)"
ws_sum["A2"].font = Font(bold=True)
ws_sum["B2"] = "=MAX(Scenarios!I2:I6)"

ws_sum["A4"] = "Scenario with Highest Portfolio Value"
ws_sum["A4"].font = Font(bold=True)
# INDEX + MATCH to find scenario
ws_sum["B4"] = "=INDEX(Scenarios!A2:A6, MATCH(B2, Scenarios!I2:I6, 0))"

ws_sum.column_dimensions["A"].width = 35
ws_sum.column_dimensions["B"].width = 25

# Save workbook
file_path = "smith_maneuver_scenario_analysis.xlsx"
wb.save(file_path)
